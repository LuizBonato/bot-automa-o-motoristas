import time
import re
import os
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

# Configuração do arquivo Excel
ARQUIVO_EXCEL = "MotoristasRegistro.xlsx"

# 🏗️ Garante que o arquivo exista com cabeçalhos
def criar_ou_encontrar_linha(telefone):
    # Abre o arquivo
    wb = load_workbook(ARQUIVO_EXCEL)
    ws = wb["Cadastro"]
    
    # Procura o telefone na planilha
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[2]) == telefone:
            wb.close()
            return i

    # Se não encontrar, cria uma nova linha
    nova = ["", "", telefone, "Primeiro contato", "", "", "", "", datetime.now().strftime("%d/%m/%Y")]
    ws.append(nova)
    wb.save(ARQUIVO_EXCEL)
    wb.close()
    return ws.max_row

# Função para salvar na planilha
def salvar_na_planilha(dados, completo=True):
    # Abre o arquivo
    wb = load_workbook(ARQUIVO_EXCEL)
    aba = "Cadastros Completos" if completo else "Contatos Incompletos"
    ws = wb[aba]

    # Encontrar a linha
    telefone = dados.get("Telefone", "")
    linha = criar_ou_encontrar_linha(telefone)

    # Preencher os campos
    campos = ["Nome", "CPF", "Telefone", "Região", "Agregado", "Veículo", "Placa", "Data Cadastro"]
    for i, campo in enumerate(campos):
        if campo in dados:
            ws.cell(row=linha, column=i+1, value=dados[campo])

    wb.save(ARQUIVO_EXCEL)
    wb.close()

# 🧠 Função para extrair dados da mensagem
def extrair_dados(mensagem):
    campos = {
        "Nome": r"Nome:\s*(.*?)(?=\n|$)",
        "CPF": r"CPF:\s*([\d\.\-]+)",
        "Telefone": r"Telefone:\s*(\(?\d+\)?[\s\-]?\d+[\-]?\d+)",
        "Região": r"Região:\s*(.*?)(?=\n|$)",
        "Agregado": r"Agregado:\s*(.*?)(?=\n|$)",
        "Veículo": r"Veículo:\s*(.*?)(?=\n|$)",
        "Placa": r"Placa:\s*([A-Za-z0-9]+)"
    }
    
    dados = {}
    for campo, padrao in campos.items():
        match = re.search(padrao, mensagem, re.IGNORECASE | re.MULTILINE)
        if match:
            valor = match.group(1).strip()
            if valor:  # Só adiciona se não estiver vazio
                dados[campo] = valor
            else:
                dados[campo] = ""  # Caso não tenha dado, deixamos como string vazia
    
    return dados

# 🕒 Configura o Chrome e abre o WhatsApp Web
chrome_options = Options()
chrome_options.add_argument(r"user-data-dir=C:\\Users\\BLD Logistica\\AppData\\Local\\Google\\Chrome\\User Data")
chrome_options.add_argument("profile-directory=Profile 7")

driver = webdriver.Chrome(options=chrome_options)
driver.get("https://web.whatsapp.com")
print("🕒 Aguardando login no WhatsApp Web...")
time.sleep(15)  # Tempo para login manual

# 🧠 Função para verificar as mensagens e extrair dados
def monitorar_mensagens():
    mensagens_lidas = set()
    print("\n🟢 Bot rodando... aguardando mensagens")
    
    try:
        while True:
            time.sleep(2)
            mensagens = driver.find_elements(By.CSS_SELECTOR, "div.message-in, div.message-out")

            if not mensagens:
                continue

            try:
                contato = driver.find_element(By.CSS_SELECTOR, "header span[title]").get_attribute("title").strip()
            except:
                continue

            ultima_msg = mensagens[-1].text.strip()

            if not ultima_msg or ultima_msg in mensagens_lidas:
                continue

            telefone_limpo = re.sub(r'\D', '', contato) if re.fullmatch(r"(\+?\d[\d\s\-().]+)", contato) else ""
            
            dados = extrair_dados(ultima_msg)
            salvar_na_planilha(dados)

            mensagens_lidas.add(ultima_msg)
            print(f"✅ Processado: {contato} | Mensagem: {ultima_msg}")

    except KeyboardInterrupt:
        print("\n🛑 Bot encerrado manualmente.")
        driver.quit()

# Executando o monitoramento
monitorar_mensagens()
